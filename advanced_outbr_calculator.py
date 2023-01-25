import pandas as pd
import datetime
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
import os

rates,rates_dates = [1],[0]

def load_file(file_path):
    df = pd.read_csv(file_path, encoding = "ISO-8859-1")
    return df.fillna(0)

def split_payment(payment,interest_balance, penalty_balance, principal_balance):
    if payment == 0:
        return (0,0,0,0)
    cash = payment
    if (cash)<penalty_balance:
        return (cash,0,0,0)
    cash-=penalty_balance
    
    if cash<interest_balance:
        return (penalty_balance,cash,0,0)
    cash-=interest_balance
    
    if cash<principal_balance:
        return (penalty_balance,interest_balance,cash,0)
    cash-=principal_balance
    
    return (penalty_balance,interest_balance,principal_balance,cash)

def get_rate(date):
    i = 0
    try:
        rates_dates[i]
    except:
        return 0
    while (rates_dates[i])<=date:
        i+=1
        if i==len(rates_dates):
            return rates[-1]
    return rates[i-1]

def get_payments(dates,usd_payments,zwl_payments,from_date,to_date):
    inbetween_dates = []
    inbetween_usd_payments = []
    inbetween_zwl_payments = []
    exchange_rates = []

    for i,date in zip(range(len(dates)),dates):
        if int(date)>=from_date and int(date)<to_date:
            if zwl_payments[i]>0:
                inbetween_dates.append(int(date))
                exc_rate = get_rate(int(date))
                inbetween_zwl_payments.append(zwl_payments[i])
                exchange_rates.append(exc_rate)
                inbetween_usd_payments.append(zwl_payments[i]/exc_rate)
            elif usd_payments[i]>0:
                inbetween_dates.append(int(date))
                inbetween_zwl_payments.append(0)
                exchange_rates.append(0)
                inbetween_usd_payments.append(usd_payments[i])
    inbetween_usd_payments = [x for _, x in sorted(zip(inbetween_dates,inbetween_usd_payments))]
    inbetween_zwl_payments = [x for _, x in sorted(zip(inbetween_dates,inbetween_zwl_payments))]
    exchange_rates = [x for _, x in sorted(zip(inbetween_dates,exchange_rates))]
    inbetween_dates.sort()
    return inbetween_dates,inbetween_usd_payments,inbetween_zwl_payments,exchange_rates

def datify(d):
    return (datetime.date(1900,1,1)+datetime.timedelta(d))

def calculate_balances(type,tenure,interest_amount,principal_amount,penalty_rate,payment_dates,usd_payments,zwl_payments,interest_rate,due_dates,grace_period,in_duplum_rule):
    intrest_added       = [0]
    penalty_added       = [0]
    principal_added     = [0]
    intrest_balance     = [0]
    penalty_balance     = [0]
    principal_balance_1 = [0]
    principal_balance_2 = [principal_amount*(tenure-grace_period)]
    loan_payment        = [0]
    loan_payment_zwl    = [0]
    dates               = [datify(due_dates[0])]
    extra_payment       = 0
    exc_rate            = [0]
    payment_to_penalty  = [0]
    payment_to_interest = [0]
    payment_to_principal= [0]
    excess_payments     = [0]
    total_interest      = 0
    total_penalty       = 0
    in_duplum           = False

    for t in range(len(due_dates)-1):
        date = (datify(due_dates[t]))
        if t%2 != 0:
            if t<tenure*2:
                new_interest_amount = interest_amount
                new_principal_amount = 0
                if t>grace_period*2:
                    new_principal_amount = principal_amount
            else:
                new_interest_amount = interest_rate*principal_balance_1[-1]
                new_principal_amount = 0
            if in_duplum_rule and t>=tenure*2:
                if in_duplum:
                    new_interest_amount = 0
                elif penalty_balance[-1]+intrest_balance[-1]+new_interest_amount >= principal_balance_2[-1]:
                    in_duplum = True
                    new_interest_amount = max(principal_balance_2[-1]-penalty_balance[-1]-intrest_balance[-1],0)
            
            total_interest +=new_interest_amount
            new_interest_balance = intrest_balance[-1]+new_interest_amount
            new_principal_balance = principal_balance_1[-1]+new_principal_amount
            
            intrest_added.append(new_interest_amount)
            intrest_balance.append(max(new_interest_balance,0))
            principal_added.append(0)
            principal_balance_1.append(principal_balance_1[-1])
            
            principal_added.append(new_principal_amount)
            principal_balance_1.append((new_principal_balance))
            intrest_added.append(0)
            intrest_balance.append(intrest_balance[-1])
            
            for i in range(2):
                penalty_added.append(0)
                penalty_balance.append(penalty_balance[-1])
                principal_balance_2.append(principal_balance_2[-1])
                loan_payment.append(0)
                loan_payment_zwl.append(0)
                dates.append(date)
                exc_rate.append(0)
                payment_to_penalty.append(0)
                payment_to_interest.append(0)
                payment_to_principal.append(0)
                excess_payments.append(0)
        else:
            new_penalty = 0
            if t>grace_period*2:
                new_penalty = (principal_balance_1[-1]+intrest_balance[-1]+penalty_balance[-1])*penalty_rate
            if in_duplum_rule:
                if in_duplum:
                    new_penalty = 0
                elif penalty_balance[-1]+intrest_balance[-1]+new_penalty >= principal_balance_2[-1]:
                    in_duplum = True
                    new_penalty = principal_balance_2[-1]-penalty_balance[-1]-intrest_balance[-1]
            total_penalty += new_penalty
            new_penalty_balance =  penalty_balance[-1]+new_penalty
            penalty_added.append(new_penalty)
            penalty_balance.append(max(new_penalty_balance,0))
            intrest_added.append(0)
            intrest_balance.append(intrest_balance[-1])
            principal_added.append(0)
            principal_balance_1.append(principal_balance_1[-1])
            principal_balance_2.append(principal_balance_2[-1])
            loan_payment.append(0)
            loan_payment_zwl.append(0)
            dates.append(date)
            exc_rate.append(0)
            payment_to_penalty.append(0)
            payment_to_interest.append(0)
            payment_to_principal.append(0)
            excess_payments.append(0)
        inbetween_dates,inbetween_usd_payments,inbetween_zwl_payments,exchange_rates = get_payments(payment_dates,usd_payments,zwl_payments,due_dates[t],due_dates[t+1])
        if len(inbetween_usd_payments)>0:
            inbetween_usd_payments[0] = inbetween_usd_payments[0]+extra_payment
        else:
            inbetween_dates.append(due_dates[t])
            inbetween_usd_payments.append(extra_payment)
            inbetween_zwl_payments.append(0)
            exchange_rates.append(0)
        
        for i in range(len(inbetween_dates)):
            total_payment = inbetween_usd_payments[i]
            penalty_payment, interest_payment, principal_payment,extra_payment = split_payment(total_payment,intrest_balance[-1],penalty_balance[-1],principal_balance_1[-1])
            if penalty_payment + interest_payment + principal_payment<0.01 and t != 0 and inbetween_zwl_payments[i]==0:
                pass
            else:
                payment_to_penalty.append(penalty_payment)
                payment_to_interest.append(interest_payment)
                payment_to_principal.append(principal_payment)
                excess_payments.append(extra_payment)
                loan_payment.append(inbetween_usd_payments[i])
                loan_payment_zwl.append(inbetween_zwl_payments[i])
                exc_rate.append(exchange_rates[i])
                principal_balance_1.append(principal_balance_1[-1]-principal_payment)
                principal_balance_2.append(principal_balance_2[-1]-principal_payment)
                intrest_balance.append(intrest_balance[-1]-interest_payment)
                penalty_balance.append(penalty_balance[-1]-penalty_payment)
                penalty_added.append(0)
                intrest_added.append(0)
                principal_added.append(0)
                dates.append(datify(inbetween_dates[i]))

        if principal_balance_2[-1]<=0.05:
            break
    if type==1:
        return (dates,penalty_added,intrest_added,principal_added,penalty_balance,intrest_balance,principal_balance_1,principal_balance_2,loan_payment,
            loan_payment_zwl,exc_rate,payment_to_penalty,payment_to_interest,payment_to_principal,excess_payments)
    else:
        return [penalty_balance[-1],intrest_balance[-1],max(0,principal_balance_1[-1]),max(0,principal_balance_2[-1]),penalty_balance[-1]+intrest_balance[-1]+max(0,principal_balance_2[-1])]

   
def collect_information(data, row):
    disbursement_date = int(data.iloc[[row],6])-2
    paid_off = int(data.iloc[[row],3])
    tenure = int(data.iloc[[row],7])
    interest_amount = (data.iloc[[row],5]).values[0]
    principal_amount= (data.iloc[[row],4]).values[0]
    interest_rate = (data.iloc[[row],8]).values[0]
    payments = list(data.iloc[[row],10:].values[0])
    penalty_rate= (data.iloc[[row],9]).values[0]
    payments_dates =(list(data.columns[10:]))
    return (tenure,interest_amount,principal_amount,interest_rate,payments_dates,payments,penalty_rate,paid_off,disbursement_date)

def save_data(df, save_path, file_name):
    try:
        df.to_csv(save_path)
    except:
        print("File could not be saved because it is open. Close the file and restart.")
        return ('Close the file {} and try again'.format(file_name),False,"")
    else:
        print("Done")
        return ('Saved successfully to {}'.format(file_name),True,save_path)

def save_xlsx(wb ,save_path, name):
    try:
        wb.save(save_path)
    except:
        print("File could not be saved because it is open. Close the file and restart.")
        return ('Close the file {} and try again'.format(name),False,"")
    else:
        print("Done")
        return ('Saved successfully to {}'.format(name),True,save_path)

def date():
    dates = []
    for year in [2020, 2021, 2022]:
        for month in ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]:
            dates.append("{}-{}".format(month,year))
    return dates

def dates_in_month(days):
    days_per_month = [31,28,31,30,31,30,31,31,30,31,30,31]
    date = datify(days) 
    month = date.month-1
    year = date.year
    if month==2:
        if year%4==0:
            return 29
    return days_per_month[month]

def get_dates_due(disbursement_date,final_date,tenure,add_interest_not_yet_due,add_penalty):
    penalty_days  = 5
    due_days      = [int(disbursement_date)]
    term_lapsed   = 0
    current       = int(disbursement_date)
    today         = (datetime.date.today()-datetime.date(1900,1,1)).days
    while current<final_date:
        term_lapsed+=1
        current+=dates_in_month(current)
        penalty_date = current+penalty_days
        if current < final_date:
            due_days.append(current)
        if penalty_date < final_date and add_penalty:
            due_days.append(penalty_date)
    if term_lapsed-1<tenure and add_interest_not_yet_due:
        due_days.append(max(final_date-1,today))
    due_days.append(final_date)
    return due_days

def get_payment_dates(disbursement_date,dates,final_date):
    disb_day = datify(disbursement_date).day
    # CHANGE THIS TO THE DAY OF REPAYMENT
    # disb_day = 1
    payment_dates = []
    for i,date in enumerate(dates):
        this_date = datify(int(date))
        d = disb_day
        date_found = False
        while not date_found:
            try:
                new_date = (datetime.date(this_date.year,this_date.month,d)-datetime.date(1900,1,1)).days
            except:
                d-=1
            else:
                date_found = True
            # To allow payments in that months that have not yet reached the instlment date to show
            if i==len(dates)-1 and new_date>=final_date:
                new_date = final_date-1
        payment_dates.append(new_date)
    return payment_dates

def normalise_dates(payment_dates):
    r = []
    for i in payment_dates:
        r.append(int(i)-2)
    return r

def get_data_1(data,loan_index):
    return collect_information(data, loan_index)

def get_data_2(data,index):
    account_id  = (data.iloc[[index],0]).values[0]
    tenure = (data.iloc[[index],4]).values[0]
    loan_amount = (data.iloc[[index],3]).values[0]
    instalment = (data.iloc[[index],7]).values[0]
    grace_period = (data.iloc[[index],9]).values[0]
    principal_amount = loan_amount/(tenure-grace_period)

    if( grace_period>0):
        interest_amount = (instalment-loan_amount)/tenure
    else:
        interest_amount = instalment - principal_amount
    interest_rate = (data.iloc[[index],5]).values[0]
    penalty_rate =(data.iloc[[index],6]).values[0]
    disbursement_date =(data.iloc[[index],8]).values[0]-2
    zwl_loan_amount = (data.iloc[[index],2]).values[0]
    return account_id,tenure,interest_amount,principal_amount,interest_rate,penalty_rate,disbursement_date,zwl_loan_amount,grace_period


def get_outbr_balances(data,final_date, path, product_prefix):
    all = []
    if product_prefix=="sl-":
        payments_data = load_file(path+'payments.csv')
    for row in range(len(data.index)):
    # for k in range(1):
    #     row = 174
        tenure,interest_amount,principal_amount,interest_rate,all_payment_dates,payments,penalty_rate,paid_off,disbursement_date = collect_information(data, row)
        due_dates = get_dates_due(disbursement_date,final_date,tenure, False,True)

        
        payment_dates = (get_payment_dates(disbursement_date,all_payment_dates,final_date))
        
        if product_prefix == "sl-":
            sl_id = data.iloc[[row],1].values[0]
            sl_payments = get_sl_payments(payments_data, sl_id)
            payment_dates = sl_payments[1]
            payments  = sl_payments[0]
        zwl_payments = [0]*len(payments)

        if paid_off==1:
            l = [(data.iloc[[row],0]).values[0],(data.iloc[[row],2]).values[0]]+[0,0,0,0,0]
        else:
            l = [(data.iloc[[row],0]).values[0],(data.iloc[[row],2]).values[0]]+calculate_balances(2,tenure,interest_amount,principal_amount,penalty_rate,payment_dates,payments,zwl_payments,interest_rate,due_dates,0,True)
        all.append(l)
        new_balances = pd.DataFrame(all,columns =['Account id', 'Account Name', 'Penalty Balance','Interest in Arrears', 'Principal in arrears',
                                'Loan Balance', 'Outstanding Balance @{}'.format(datify(int(final_date-1)))])
    return new_balances

def create_statement(tenure,interest_amount,principal_amount,interest_rate,payment_dates,usd_payments,zwl_payments,penalty_rate,disbursement_date,frequency, final_date,grace_period,in_duplum_rule):
    due_dates = get_dates_due(disbursement_date,final_date,tenure, True,True)
    

    (dates,penalty_added,intrest_added,principal_added,penalty_balance,intrest_balance,principal_balance_1,principal_balance_2,loan_payment,
            loan_payment_zwl,exc_rate,payment_to_penalty,payment_to_interest,payment_to_principal,excess_payments) = calculate_balances(1,tenure,interest_amount,principal_amount,penalty_rate,payment_dates,usd_payments,zwl_payments,interest_rate,due_dates,grace_period,in_duplum_rule)
    entries = []
    dates_index = []
    for i in range(len(dates)):
        description = ""
        dr = penalty_added[i]+intrest_added[i]+principal_added[i]
        cr = payment_to_penalty[i]+payment_to_interest[i]+payment_to_principal[i]+excess_payments[i]
        if i==0:
            description="Loan Disbursement"
        elif penalty_added[i]>0:
            description = "Penalty Due"
        elif intrest_added[i]>0:
            description = "Interest Due"
        elif principal_added[i]>0:
            description = "Principal Due"
        elif cr>0:
            description = "Loan Repayment"
        else:
            continue
        dates_index.append(dates[i])
        entries.append([description,dr,cr,penalty_balance[i],intrest_balance[i],principal_balance_1[i],principal_balance_2[i],
                    loan_payment_zwl[i],exc_rate[i],loan_payment[i],payment_to_penalty[i],payment_to_interest[i],payment_to_principal[i],
                    excess_payments[i]])
    new_balances = pd.DataFrame(entries,columns =['Description', 'DR', 'CR','Penalty balance', 'Interest Balance',
                                'Principal Balance', 'Loan Balance','Payment (ZWL)', 'Exchange Rate','Usd Payments', "Penalty Payment",
                                "Interest Payment","Principal Payment", "Extra"],index=dates_index)
    return new_balances

def get_information(name,tenure,zwl_loan_amount,interest_amount,principal_amount,interest_rate,final_date,payments,penalty_rate,convert,outstanding_balance):
    last_rate = 0
    if convert:
        last_rate = rates[-1]
    return {"Statement Date": (datify(final_date-1)),
        "Client Name": (name),
        "Loan amount ZWL":(zwl_loan_amount),
        "Loan amount USD": (principal_amount*tenure),
        "Tenure": (tenure),
        "Instalment": (principal_amount+interest_amount),
        "Instalment Principal" : (round(principal_amount*100)/100),
        "Instalment Interest": (round(interest_amount*100)/100),
        "Interest Rate":(interest_rate),
        "Penalty Rate":(penalty_rate),
        "Total repayments":(sum(payments)),
        "Outstanding balance USD":(outstanding_balance),
        "Current Rate":(last_rate),
        "Outstanding balance ZWL":(outstanding_balance*last_rate)}

def add_details(info_dict,wb,sheet):
    current_sheet = wb[sheet]
    keys = list(info_dict.keys())
    values = list(info_dict.values())
    for i in range(len(keys)):
        current_sheet.cell(row =  i+1, column=1, value=keys[i])
        current_sheet.cell(row =  i+1, column=2, value=values[i])
    return wb

def dataframe_to_xlsx(df,wb,sheet,start_row,start_col):
    current_sheet = wb[sheet]
    rows = dataframe_to_rows(df)
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            current_sheet.cell(row=r_idx+start_row, column=c_idx+start_col, value=value)
    return wb

# def get_sl_payments(path,loan_index):
#     file_path = path+"payments.csv"
#     df = pd.read_csv(file_path)
#     paid_off = int(df.iloc[[loan_index],3])
#     rates_dates =normalise_dates(df["Dates"].tolist())

def download_payments_from_file(payments_data,extra,account_id):
    payments = []
    dates    = []
    if not payments_data.empty:
        main_payments_filtered  = payments_data.loc[payments_data['Account ID'] == account_id]
        dates                  += main_payments_filtered['Trx Date'].tolist()
        payments               += main_payments_filtered['Trx Amount'].tolist()
    if not extra.empty:
        extra_payments_filtered = extra.loc[extra['Account ID'] == account_id]
        dates                  += extra_payments_filtered['Trx Date'].tolist()
        payments               += extra_payments_filtered['Trx Amount'].tolist()
    return (payments,dates)

def reset_rates():
    global rates,rates_dates
    rates,rates_dates = [1],[0]
    
def load_rates(start_path):
    file_path = start_path+"Payoffs\\Statements\\sources\\Rates.csv"
    df = pd.read_csv(file_path)
    global rates,rates_dates
    rates = df["Rates"].tolist()
    rates_dates =normalise_dates(df["Dates"].tolist())

def get_sl_payments(payments_data,account_id):
    payments  = []
    dates     = []
    deposits  = []
    lamp_type = []
    main_payments_filtered  = payments_data.loc[payments_data['Account ID'] == account_id]
    dates                  += main_payments_filtered['Trx Date'].tolist()
    deposits               += main_payments_filtered['Down Payment'].tolist()
    lamp_type              += main_payments_filtered['Group Name'].tolist()
    payments               += main_payments_filtered['Trx Amount'].tolist()
    
    final_dates   = []
    final_amounts = []
    for i in range(len(dates)):
        paid_amount = payments[i]
        if deposits[i]==True:
            if lamp_type[i] == "Sunking Home 60 Live":
                paid_amount = payments[i]-30
                if paid_amount==0:
                    continue
            elif lamp_type[i] == "Sunking Home 120 Live (Credit)":
                paid_amount = payments[i]-45
                if paid_amount==0:
                    continue
        final_dates.append(dates[i])
        final_amounts.append(paid_amount)
    return (final_amounts,normalise_dates(final_dates))

